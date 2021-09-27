'''
select series (T1+C, T2 FLAIR, ASL, Standard-dose PET) from the raw dataset
copy the selected series to another folder
'''

import numpy as np
import os
from glob import glob
import dicom
import pdb
import openpyxl as px
import shutil

src_data_path = '../../data/GBM/'
dst_data_path = '../../data/GBM_selected/'
subj_info_path = '../../data/GBM.xlsx'

if not os.path.exists(dst_data_path):
    os.makedirs(dst_data_path)

# subj_list = ['Fdg_Stanford_001', 'Fdg_Stanford_003', 'Fdg_Stanford_004', 'Fdg_Stanford_005',
#             'Fdg_Stanford_010', 'Fdg_Stanford_011', 'Fdg_Stanford_012', 'Fdg_Stanford_014',
#             'Fdg_Stanford_015', 'Fdg_Stanford_017', '852_06182015', '869_06252015',
#             '1496_05272016', '1498_05312016', '1512_06062016', '1549_06232016',
#             '1559_06282016', '1572_07052016', '1604_07142016', '1610_07152016',
#             '1619_07192016', '2002_02142017', '2010_02212017', '2120_04122017',
#             '2142_04242017', '2275_07062017', '2277_07072017', '2284_07112017',
#             '2292_07122017', '2295_07122017', '2310_07242017', '2374_08242017',
#             '2377_08252017', '2388_08302017']
# ['842_06152015', '15632_06042015', '1657_08022016', '1657_08022016']

subj_num = 34

subj_info_xls = px.load_workbook(subj_info_path)
sheet = subj_info_xls.get_sheet_by_name(name='series')

pdb.set_trace()

for i in range(2, 2+subj_num):
    subj_name = str(sheet.cell(column=1, row=i).value)
    subj_name_path = src_data_path + subj_name + '/'
    print('Starting '+subj_name)
    if not os.path.exists(subj_name_path):  # no this subject
        print(subj_name+' does not exist!')
        continue
    T1_C_idx = sheet.cell(column=6, row=i).value
    T2_FLAIR_idx = sheet.cell(column=8, row=i).value
    ASL_idx = sheet.cell(column=10, row=i).value
    PET_idx = sheet.cell(column=12, row=i).value
    if T1_C_idx == None or T2_FLAIR_idx == None or ASL_idx == None or PET_idx == None:
        print(subj_name+' missing sequence!')
        continue

    # find a valid subject scanning
    dst_name_path = dst_data_path + subj_name + '/'
    if not os.path.exists(dst_name_path):
        os.makedirs(dst_name_path)
    # series_dict = {'T1':int(T1_C_idx), 'T2_FLAIR':int(T2_FLAIR_idx),
    #                 'ASL':int(ASL_idx), 'PET'int(PET_idx)}
    series_dict = {int(T1_C_idx):'T1', int(T2_FLAIR_idx):'T2_FLAIR',
                    int(ASL_idx):'ASL', int(PET_idx):'PET'}
    series_paths = glob(subj_name_path+'*')
    if len(series_paths) == 1:        # go to the sub-folder if only one sub-folder
        series_paths = glob(series_paths[0]+'/*')
    for series_path in series_paths:    # go through each folder to check whether is needed
        file_paths = glob(series_path+'/*')
        if len(file_paths) == 0:        # no file inside
            continue
        dcm = dicom.read_file(file_paths[0])
        dcm_series_num = int(dcm.SeriesNumber)
        dcm_series_modality = dcm.Modality
        print(dcm_series_num)
        print(dcm_series_modality)
        if dcm_series_num in series_dict:   # check series ID
            # pdb.set_trace()
            modality = 'PT' if series_dict[dcm_series_num] == 'PET' else 'MR'   # check modality
            if dcm_series_modality == modality:
                dst_series_path = dst_name_path + series_dict[dcm_series_num] + '/'
                if os.path.exists(dst_series_path):
                    shutil.rmtree(dst_series_path)
                shutil.copytree(series_path, dst_series_path)
